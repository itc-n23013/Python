import openpyxl
import sys

if len(sys.argv) < 2:
    sys.exit('Usage: python swaprowcol.py example.xlsx')

src = sys.argv[1]

wb = openpyxl.load_workbook(src)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        old_cell = sheet.cell(column=col, row=row)

        new_cell = new_sheet.cell(column=row, row=col)
        new_cell.value = old_cell.value

new_wb.save(src + '.swap.xlsx')
